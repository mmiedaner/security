import numpy
import openpyxl
import pandas


class ExcelAdaptor:
    """
    ExcelAdaptor to handle all operations regarding excel files
    """

    def __find_empty_column(self, sheet):
        """
        Finds empty column in sheet
        :param sheet: sheet to analyse
        :return: column
        """
        column = 1;
        while sheet.cell(row=1, column=column).value is not None:
            column += 1
        return column

    def append_data_to_excelsheet(self, workbook, file_name, sheet_name, column_title, data):
        """
        Update excel sheet with single columned data set
        :param file_name: excel workbook to update
        :param sheet_name: excel sheet to update
        :param column_title: title of column to be set on empty column
        :param data: single row dataframe containing all data
        :return: None
        """
        if workbook is None:
            workbook = openpyxl.load_workbook(file_name)

        sheet = workbook.get_sheet_by_name(sheet_name)

        column = self.__find_empty_column(sheet)
        values = data.values
        row = 2
        sheet.cell(row=1, column=column).value = column_title
        for value in numpy.nditer(values):
            sheet.cell(row=row, column=column).value = int(value)
            row += 1

        return workbook

    def write_all_changes_to_excel(self, workbook, file_name):
        workbook.save(file_name)

    def create_pivot_analysis(self, df_togroup, column_name_to_group, helper_column, set_of_bins, set_of_labels,
                              label_header):
        """
        Generates pivot table
        :param df_togroup: dataframe with data to be analyzed
        :param column_name_to_group: name of column to group data by
        :param helper_column: name of helper column
        :param set_of_bins: set of bins to group data into
        :param set_of_labels: set of labels to name bins
        :param label_header: label header
        :return:
        """
        labels_to_use = numpy.asarray(set_of_labels)
        bins_to_use = numpy.asarray(set_of_bins)
        df_toreturn = df_togroup.groupby(
            pandas.cut(df_togroup[column_name_to_group], bins=bins_to_use, labels=labels_to_use)).count()
        for name in list(df_toreturn):
            if name != column_name_to_group and name != helper_column:
                del df_toreturn[name]

        series_items = numpy.arange(125, 750, 125)
        df_toreturn[label_header] = pandas.Series(series_items, index=df_toreturn.index)

        return df_toreturn

    def write_to_singleframe_to_excel(self, dataframe, filename, sheetname):
        """
        Writes dataframe to excel sheet_static
        :param filename: name of excel to write to
        :param sheetname: name of sheet_static to write into
        :return:
        """
        writer = pandas.ExcelWriter(filename)
        dataframe.to_execl(writer, sheet_name=sheetname, index=False)
        writer.save()

    def write_multiple_frames_to_one_sheet(self, dataframes, filename, sheetnames):
        """
        Writes multiple data frames into one sheet_static.
        :param dataframes:
        :param filename:
        :param sheetnames:
        :return:
        """
        writer = pandas.ExcelWriter(filename)
        i = 0
        for dataframe in dataframes:
            dataframe.to_excel(writer, sheet_name=sheetnames[i], index=False)
            i += 1

        writer.save()

    def read_sheet_from_excel(self, filename, sheetname):
        """
        Reads content of excel sheet_static
        :param filename: name of file to read from
        :param sheetname: name of sheet_static to read into a dataframe
        :return: dataframe read from excel
        """
        excel_input = pandas.ExcelFile(filename).parse(sheetname)
        return excel_input